import os
from cs50 import SQL
from flask import Flask, flash, redirect, render_template, request, session,url_for
from flask_session import Session
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime
from helpers import apology, login_required, lookup, get_translations
from flask import send_from_directory
from flask import request, jsonify
from dropdowns import reasons, department, action, priority, status, smth, talep  # these are just for getting the dropdown lists
import requests
import re
from collections import defaultdict
from openpyxl import load_workbook
# ------------------------
# Configure application
# ------------------------
app = Flask(__name__)




# serve files under static/files/uploads/<folder>/<...>
@app.route('/static/files/uploads/<path:filename>')
def uploaded_file(filename):
    # Use app.static_folder, and path relative to it
    uploads_root = os.path.join(app.static_folder, "files", "uploads")
    # send the correct subpath
    return send_from_directory(uploads_root, filename)


BASE_DIR = os.path.dirname(os.path.dirname(__file__))  # backend/..
db_path = os.path.join(BASE_DIR, "/static/files", "unimak.db")

# Configure session
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

# Connect to database
db = SQL("sqlite:///static/files/unimak.db")

# ------------------------
# After request: prevent caching
# ------------------------
@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Expires"] = 0
    response.headers["Pragma"] = "no-cache"
    return response

# -------------------- HOME --------------------
@app.route("/", methods=["GET"])
@login_required
def index():
    problems = db.execute("""
        SELECT 
            p.id AS problem_id,
            p.created_at,
            pj.project_number,
            pj.project_name,
            m.manager_name,
            c.customer_name,
            g.group_name,
            g.group_number,
            e.engineer_name,
            pc.id AS pc_id,
            comp.component_name,
            comp.component_no,
            pc.reason,
            pc.description,
            pc.priority,
            pc.action,
            pc.department,
            ps.df_filename
        FROM problems p
        JOIN projects pj ON pj.id = p.project_id
        JOIN managers m ON pj.manager_id = m.id
        JOIN customers c ON pj.customer_id = c.id
        JOIN groups g ON g.id = p.group_id
        JOIN engineers e ON g.engineer_id = e.id
        LEFT JOIN problem_components pc ON pc.problem_id = p.id
        LEFT JOIN components comp ON comp.id = pc.component_id
        LEFT JOIN problem_steps ps ON ps.problem_id = p.id
        ORDER BY p.created_at DESC
    """)

    problem_dict = defaultdict(lambda: {"components": [], "photos": []})

    for row in problems:
        prob = problem_dict[row["problem_id"]]
        prob.update({
            "problem_id": row["problem_id"],
            "created_at": row["created_at"],
            "project_number": row["project_number"],
            "project_name": row["project_name"],
            "manager_name": row["manager_name"],
            "customer_name": row["customer_name"],
            "group_name": row["group_name"],
            "group_number": row["group_number"],
            "engineer_name": row["engineer_name"],
            "df_filename": row["df_filename"]
        })
        if row["pc_id"]:
            prob["components"].append({
                "component_name": row["component_name"],
                "component_no": row["component_no"],
                "reason": row["reason"],
                "description": row["description"],
                "priority": row["priority"],
                "action": row["action"],
                "department": row["department"]
            })

    # Attach photos from filesystem
    base_path = os.path.join(app.root_path, "static/files/uploads")
    for prob in problem_dict.values():
        df_prefix = prob["df_filename"].split(".")[0] if prob["df_filename"] else None
        if df_prefix:
            pictures_dir = os.path.join(base_path, df_prefix, "pictures")
            if os.path.exists(pictures_dir):
                prob["photos"] = os.listdir(pictures_dir)

    data = list(problem_dict.values())
    t = get_translations()
    return render_template("home.html", data=data, t=t)


# -------------------- LOGIN --------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    session.clear()
    if request.method == "POST":
        if not request.form.get("username"):
            return apology("must provide username", 400)
        elif not request.form.get("password"):
            return apology("must provide password", 400)

        rows = db.execute("SELECT * FROM users WHERE username = ?", request.form.get("username"))
        if len(rows) != 1 or not check_password_hash(rows[0]["password_hash"], request.form.get("password")):
            return apology("invalid username and/or password", 400)

        session["user_id"] = rows[0]["id"]
        session["language"] = rows[0].get("language", "en")
        return redirect("/")
    else:
        return render_template("login.html")

# -------------------- LOGOUT --------------------
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# -------------------- REGISTER --------------------
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        confirmation = request.form.get("confirmation")

        if not username or not password or not confirmation:
            return apology("Please fill all fields", 400)
        if password != confirmation:
            return apology("Passwords do not match", 400)
        if db.execute("SELECT * FROM users WHERE username = ?", username):
            return apology("Username already taken", 400)

        hash_pw = generate_password_hash(password)
        db.execute("INSERT INTO users (username, password_hash) VALUES (?, ?)", username, hash_pw)
        return redirect("/login")
    else:
        return render_template("register.html")

# -------------------- UPLOAD PROBLEM --------------------
@app.route("/upload", methods=["GET", "POST"])
@login_required
def upload():
    # --- fetch dropdowns and DB references (always defined for GET and POST) ---
    managers = db.execute("SELECT id, manager_name, manager_mail FROM managers")
    customers = db.execute("SELECT id, customer_name, customer_country FROM customers")
    projects = db.execute("""
        SELECT p.id, p.project_number, p.project_name, p.quantity,
               p.manager_id, m.manager_name, c.customer_name
        FROM projects p
        JOIN managers m ON p.manager_id = m.id
        JOIN customers c ON p.customer_id = c.id
    """)
    groups = db.execute("""
        SELECT g.id, g.project_id, e.engineer_name,
               g.group_number, g.group_name
        FROM groups g
        JOIN engineers e ON g.engineer_id = e.id
    """)
    components = db.execute("""
        SELECT c.id, c.group_id, c.component_no, c.component_name,
               c.unit_quantity, c.total_quantity
        FROM components c
    """)

    # dropdown constants (from your module)
    dropdown_reasons = reasons
    dropdown_priority = priority
    dropdown_department = department
    dropdown_action = action
    dropdown_status = status
    dropdown_smth = smth
    dropdown_talep = talep

    t = get_translations()

    def parse_components_from_form(form):
        """
        Supports two shapes:
         - array-style: component_id[], reason[], department[], action[], priority[], description[]
         - nested-style: components[0][component_id], components[0][reason], ...
        Returns list of dicts:
            [ {"component_id": "...", "reason": "...", ...}, ... ]
        """
        # 1) try array-style first
        arr_component_ids = form.getlist("component_id[]") or form.getlist("component_id")
        if arr_component_ids:
            # fetch parallel arrays (safely)
            reasons_list = form.getlist("reason[]") or form.getlist("reason")
            departments = form.getlist("department[]") or form.getlist("department")
            actions = form.getlist("action[]") or form.getlist("action")
            priorities = form.getlist("priority[]") or form.getlist("priority")
            descriptions = form.getlist("description[]") or form.getlist("description")
            status = form.getlist("status[]") or form.getlist("status")
            smth = form.getlist("smth[]") or form.getlist("smth")

            n = len(arr_component_ids)
            items = []
            for i in range(n):
                items.append({
                    "component_id": arr_component_ids[i] if i < len(arr_component_ids) else None,
                    "reason": reasons_list[i] if i < len(reasons_list) else None,
                    "department": departments[i] if i < len(departments) else None,
                    "action": actions[i] if i < len(actions) else None,
                    "priority": priorities[i] if i < len(priorities) else None,
                    "description": descriptions[i] if i < len(descriptions) else None,
                    "status": status[i] if i < len(status) else None,
                    "smth": smth[i] if i < len(smth) else None,
                    "talep": talep[i] if i < len(talep) else None
                })
            return items

        # 2) try nested-style: components[0][component_id]
        nested = defaultdict(dict)
        pattern = re.compile(r"components\[(\d+)\]\[(\w+)\]")
        for key in form.keys():
            m = pattern.match(key)
            if m:
                idx = int(m.group(1))
                field = m.group(2)
                nested[idx][field] = form.get(key)
        if nested:
            # convert to ordered list
            return [ nested[i] for i in sorted(nested.keys()) ]

        # none found -> return empty
        return []

    if request.method == "POST":
        # ---- read basic problem fields ----
        project_id = request.form.get("project_id") or None
        group_id = request.form.get("group_id") or None
        planned_closing_date = request.form.get("planned_closing_date") or None

        # ---- parse components (robust) ----
        components_list = parse_components_from_form(request.form)

        # ---- files ----
        photos = request.files.getlist("photos")  # <input name="photos" multiple>

        # ---- create df number and folders ----
        timestamp = datetime.now().strftime("%d%m%y%H%M%S")
        df_number = f"df_{timestamp}"
        folder_name = df_number
        base_dir = os.path.join(app.static_folder, "files", "uploads", folder_name)
        pictures_dir = os.path.join(base_dir, "pictures")
        os.makedirs(pictures_dir, exist_ok=True)

        saved_photo_filenames = []
        for idx, photo in enumerate(photos, start=1):
            if photo and photo.filename:
                ext = os.path.splitext(photo.filename)[1] or ".jpg"
                filename = secure_filename(f"{folder_name}_{idx}{ext}")
                photo.save(os.path.join(pictures_dir, filename))
                saved_photo_filenames.append(filename)

        # ---- Insert main problems row using cs50.SQL execute (no cursor) ----
        db.execute("""
            INSERT INTO problems (project_id, group_id, user_id, planned_closing_date)
            VALUES (?, ?, ?, ?)
        """, project_id, group_id, session.get("user_id"), planned_closing_date)

        # get inserted problem id
        problem_id = db.execute("SELECT last_insert_rowid() AS id")[0]["id"]

        # ---- Insert problem_components and problem_steps per component ----
        for comp in components_list:
            comp_id = comp.get("component_id") or comp.get("component") or None
            reason_v = comp.get("reason")
            department_v = comp.get("department")
            action_v = comp.get("action")
            priority_v = comp.get("priority")
            description_v = comp.get("description")

            db.execute("""
                INSERT INTO problem_components
                (problem_id, component_id, reason, department, action, priority, description)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, problem_id, comp_id, reason_v, department_v, action_v, priority_v, description_v)

            # add initial step row for that component (adjust fields to your schema)
            db.execute("""
                INSERT INTO problem_steps
                (problem_id, component_id, step_number, df_filename, quantity, action, status, planned_closing_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, problem_id, comp_id, 1, f"{df_number}.xlsx", 0, action_v or "Initial", "Open", planned_closing_date)

        # optionally: store photo filenames somewhere (if you add a column to problems or a photos table)
        # e.g. db.execute("UPDATE problems SET photos_id = ? WHERE id = ?", ",".join(saved_photo_filenames), problem_id)

        flash("Problem reported successfully!", "success")
        return redirect("/")

    # GET → render upload.html
    # NOTE: pass the dropdown objects under expected names
    return render_template(
        "upload.html",
        data={
            "projects": projects,
            "groups": groups,
            "components": components,
            "managers": managers,
            "customers": customers
        },
        reasons=reasons,
        department=department,
        action=action,
        status = status,
        priority=priority,
        t=t,
        smth=smth,
        talep=talep
    )


# -------------------- INFO --------------------
@app.route("/info", methods=["GET"])
@login_required
def info():
    # Single query to get managers -> projects -> problems (may produce repeated project rows when multiple problems exist)
    rows = db.execute("""
        SELECT
            m.id AS manager_id,
            m.manager_name,
            p.id AS project_id,
            p.project_number,
            p.quantity,
            pr.id AS problem_id,
            pr.df_number,
            pr.reason,
            pr.description AS problem_description,
            pr.photos_id,
            pr.status,
            pr.record_date
        FROM managers m
        LEFT JOIN projects p ON p.manager_id = m.id
        LEFT JOIN problems pr ON pr.project_id = p.id
        ORDER BY m.manager_name, p.project_number, pr.record_date DESC
    """)


    # Build nested structure: { manager_name: [ { project }, { project }, ... ] }
    data = {}
    # helper map to track added projects per manager for O(1) lookup
    project_index = {}

    for row in rows:
        manager = row["manager_name"] or "Unassigned"

        if manager not in data:
            data[manager] = []
            project_index[manager] = {}

        project_id = row["project_id"]

        # If there is a project (project_id can be None when manager has no projects)
        if project_id:
            # create project entry only once per manager
            if project_id not in project_index[manager]:
                project_obj = {
                    "id": project_id,
                    "project_number": row["project_number"],
                    "quantity": row.get("quantity"),
                    "machine_type": row.get("machine_type"),
                    "machine_top_group": row.get("machine_top_group"),
                    # placeholder for project-level description (if you add one later)
                    "description": row.get("project_description") if "project_description" in row.keys() else None,
                    "problems": []
                }
                project_index[manager][project_id] = project_obj
                data[manager].append(project_obj)

            # append problem if exists
            if row["problem_id"]:
                photos = []
                if row["photos_id"]:
                    # photos stored as comma-separated filenames
                    photos = [p.strip() for p in row["photos_id"].split(",") if p.strip()]

                problem_obj = {
                    "id": row["problem_id"],
                    "df_number": row.get("df_number"),
                    "reason": row.get("reason"),
                    "description": row.get("problem_description"),
                    "photos": photos,
                    "status": row.get("status"),
                    "record_date": row.get("record_date")
                }
                project_index[manager][project_id]["problems"].append(problem_obj)

    # ensure managers with no projects still show up (data[manager] would be empty list)
    t = get_translations() if callable(globals().get("get_translations", None)) else {}

    return render_template("info.html", data=data, t=t)

# -------------------- HISTORY --------------------
@app.route("/history", methods=["GET"])
@login_required
def history():
    rows = db.execute("""
        SELECT ps.df_filename, p.project_number, m.manager_name, pr.reason, pr.description, pr.photos_id, ps.created_at
        FROM problem_steps ps
        JOIN problems pr ON ps.problem_id = pr.id
        JOIN projects p ON pr.project_id = p.id
        JOIN managers m ON p.manager_id = m.id
        ORDER BY ps.created_at DESC
    """)
    t = get_translations()
    return render_template("history.html", data=rows, t=t)

# -------------------- ADMIN --------------------

@app.route("/admin", methods=["GET", "POST"])
@login_required
def admin():
    if request.method == "POST":
        action = request.form.get("action")
        tab = request.form.get("tab", "problems")  # Default to problems tab

        # ========== PROJECTS TAB ACTIONS ==========
        if tab == "projects":
            if action == "add_project":
                project_number = request.form.get("project_number")
                project_name = request.form.get("project_name")
                manager_id = request.form.get("manager_id")
                customer_id = request.form.get("customer_id")
                quantity = request.form.get("quantity")

                if not all([project_number, project_name, manager_id, customer_id, quantity]):
                    flash("All fields are required.", "error")
                    return redirect("/admin?tab=projects")

                try:
                    quantity = int(quantity)
                except ValueError:
                    flash("Quantity must be a number.", "error")
                    return redirect("/admin?tab=projects")

                existing = db.execute("SELECT id FROM projects WHERE project_number = ?", project_number)
                if existing:
                    flash(f"Project number '{project_number}' already exists.", "error")
                    return redirect("/admin?tab=projects")

                try:
                    db.execute("""
                        INSERT INTO projects (project_number, project_name, manager_id, customer_id, quantity)
                        VALUES (?, ?, ?, ?, ?)
                    """, project_number, project_name, manager_id, customer_id, quantity)
                    flash(f"Project '{project_number}' added successfully!", "success")
                except Exception as e:
                    flash(f"Error adding project: {str(e)}", "error")

            elif action == "edit_project":
                project_id = request.form.get("project_id")
                project_number = request.form.get("project_number")
                project_name = request.form.get("project_name")
                manager_id = request.form.get("manager_id")
                customer_id = request.form.get("customer_id")
                quantity = request.form.get("quantity")

                if not all([project_id, project_number, project_name, manager_id, customer_id, quantity]):
                    flash("All fields are required.", "error")
                    return redirect("/admin?tab=projects")

                try:
                    quantity = int(quantity)
                    # Check if project number conflicts with another project
                    existing = db.execute("SELECT id FROM projects WHERE project_number = ? AND id != ?", 
                                        project_number, project_id)
                    if existing:
                        flash(f"Project number '{project_number}' already exists.", "error")
                        return redirect("/admin?tab=projects")

                    db.execute("""
                        UPDATE projects 
                        SET project_number = ?, project_name = ?, manager_id = ?, customer_id = ?, quantity = ?
                        WHERE id = ?
                    """, project_number, project_name, manager_id, customer_id, quantity, project_id)
                    flash(f"Project '{project_number}' updated successfully!", "success")
                except Exception as e:
                    flash(f"Error updating project: {str(e)}", "error")

            elif action == "delete_project":
                project_id = request.form.get("project_id")
                if project_id:
                    problems = db.execute("SELECT id FROM problems WHERE project_id = ?", project_id)
                    if problems:
                        flash("Cannot delete project with associated problems. Please delete problems first.", "error")
                    else:
                        try:
                            db.execute("DELETE FROM projects WHERE id = ?", project_id)
                            flash("Project deleted successfully!", "success")
                        except Exception as e:
                            flash(f"Error deleting project: {str(e)}", "error")

            elif action == "upload_excel":
                if 'excel_file' not in request.files:
                    flash("No file uploaded.", "error")
                    return redirect("/admin?tab=projects")

                file = request.files['excel_file']
                if file.filename == '':
                    flash("No file selected.", "error")
                    return redirect("/admin?tab=projects")

                if not file.filename.endswith(('.xlsx', '.xls')):
                    flash("Please upload an Excel file (.xlsx or .xls).", "error")
                    return redirect("/admin?tab=projects")

                try:
                    # Save temporary file
                    temp_path = os.path.join(app.static_folder, "temp", secure_filename(file.filename))
                    os.makedirs(os.path.dirname(temp_path), exist_ok=True)
                    file.save(temp_path)

                    # Load workbook
                    wb = load_workbook(temp_path)
                    ws = wb.active

                    # Get managers and customers for lookup
                    managers_dict = {m['manager_name']: m['id'] for m in db.execute("SELECT id, manager_name FROM managers")}
                    customers_dict = {f"{c['customer_name']}": c['id'] for c in db.execute("SELECT id, customer_name FROM customers")}

                    # Read rows (assuming first row is header)
                    rows_processed = 0
                    errors = []
                    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if not any(row):  # Skip empty rows
                            continue

                        try:
                            project_number = str(row[0]).strip() if row[0] else None
                            project_name = str(row[1]).strip() if row[1] else None
                            manager_name = str(row[2]).strip() if row[2] else None
                            customer_name = str(row[3]).strip() if row[3] else None
                            quantity = row[4] if row[4] else None

                            if not all([project_number, project_name, manager_name, customer_name, quantity]):
                                errors.append(f"Row {idx}: Missing required fields")
                                continue

                            manager_id = managers_dict.get(manager_name)
                            customer_id = customers_dict.get(customer_name)

                            if not manager_id:
                                errors.append(f"Row {idx}: Manager '{manager_name}' not found")
                                continue
                            if not customer_id:
                                errors.append(f"Row {idx}: Customer '{customer_name}' not found")
                                continue

                            # Check if project exists
                            existing = db.execute("SELECT id FROM projects WHERE project_number = ?", project_number)
                            if existing:
                                # Update existing
                                db.execute("""
                                    UPDATE projects 
                                    SET project_name = ?, manager_id = ?, customer_id = ?, quantity = ?
                                    WHERE project_number = ?
                                """, project_name, manager_id, customer_id, int(quantity), project_number)
                            else:
                                # Insert new
                                db.execute("""
                                    INSERT INTO projects (project_number, project_name, manager_id, customer_id, quantity)
                                    VALUES (?, ?, ?, ?, ?)
                                """, project_number, project_name, manager_id, customer_id, int(quantity))

                            rows_processed += 1
                        except Exception as e:
                            errors.append(f"Row {idx}: {str(e)}")

                    # Clean up temp file
                    os.remove(temp_path)

                    if errors:
                        flash(f"Processed {rows_processed} rows. Errors: {'; '.join(errors[:5])}", "error")
                    else:
                        flash(f"Successfully processed {rows_processed} projects from Excel!", "success")

                except Exception as e:
                    flash(f"Error processing Excel file: {str(e)}", "error")

            return redirect("/admin?tab=projects")

        # ========== PROBLEMS TAB ACTIONS ==========
        elif tab == "problems":
            if action == "delete_problem":
                pid = request.form.get("problem_id")
                db.execute("DELETE FROM problems WHERE id = ?", pid)
                flash(f"Problem {pid} deleted.", "success")

            elif action == "update_date":
                pid = request.form.get("problem_id")
                new_date = request.form.get("planned_closing_date")
                db.execute(
                    "UPDATE problems SET planned_closing_date = ? WHERE id = ?",
                    new_date, pid
                )
                flash(f"Updated closing date for Problem {pid}.", "success")

            elif action == "update_step_status":
                step_id = request.form.get("step_id")
                new_status = request.form.get("status")
                if step_id and new_status:
                    db.execute("UPDATE problem_steps SET status = ? WHERE id = ?", new_status, step_id)
                    flash("Step status updated successfully!", "success")

            elif action == "delete_component":
                cid = request.form.get("component_id")
                db.execute("DELETE FROM problem_components WHERE id = ?", cid)
                flash(f"Component {cid} deleted.", "success")

            elif action == "delete_step":
                sid = request.form.get("step_id")
                db.execute("DELETE FROM problem_steps WHERE id = ?", sid)
                flash(f"Step {sid} deleted.", "success")

            return redirect("/admin?tab=problems")

    # ========== GET REQUEST - LOAD DATA ==========
    active_tab = request.args.get("tab", "problems")

    # Load projects data with groups and components (for Tab 1)
    projects_raw = db.execute("""
        SELECT p.id, p.project_number, p.project_name, p.quantity,
               m.id as manager_id, m.manager_name, 
               c.id as customer_id, c.customer_name, c.customer_country,
               COUNT(DISTINCT pr.id) as problem_count
        FROM projects p
        JOIN managers m ON p.manager_id = m.id
        JOIN customers c ON p.customer_id = c.id
        LEFT JOIN problems pr ON pr.project_id = p.id
        GROUP BY p.id
        ORDER BY p.project_number
    """)
    
    # Build hierarchical structure: projects -> groups -> components
    projects = []
    for proj in projects_raw:
        # Get groups for this project
        groups_raw = db.execute("""
            SELECT g.id, g.group_number, g.group_name, e.engineer_name, e.id as engineer_id
            FROM groups g
            JOIN engineers e ON g.engineer_id = e.id
            WHERE g.project_id = ?
            ORDER BY g.group_number
        """, proj["id"])
        
        groups = []
        for grp in groups_raw:
            # Get components for this group
            components = db.execute("""
                SELECT id, position_no, component_no, component_name, unit_quantity, 
                       total_quantity, weight, description, size, materials, 
                       machine_type, notes, working_area
                FROM components
                WHERE group_id = ?
                ORDER BY component_no
            """, grp["id"])
            
            groups.append({
                **grp,
                "components": components
            })
        
        projects.append({
            **proj,
            "groups": groups
        })

    managers = db.execute("SELECT id, manager_name FROM managers ORDER BY manager_name")
    customers = db.execute("SELECT id, customer_name, customer_country FROM customers ORDER BY customer_name")

    # Load problems data (for Tab 2)
    problems = db.execute("""
        SELECT p.id, p.created_at, p.planned_closing_date,
               pr.project_number, pr.project_name,
               g.group_name,
               u.username
        FROM problems p
        JOIN projects pr ON p.project_id = pr.id
        JOIN groups g ON p.group_id = g.id
        JOIN users u ON p.user_id = u.id
        ORDER BY p.created_at DESC
    """)

    reports = []
    for prob in problems:
        components = db.execute("""
            SELECT pc.id, c.component_name, pc.reason, pc.priority,
                   pc.department, pc.action
            FROM problem_components pc
            JOIN components c ON pc.component_id = c.id
            WHERE pc.problem_id = ?
        """, prob["id"])

        steps = db.execute("""
            SELECT id, step_number, df_filename, status, action, problem_id
            FROM problem_steps
            WHERE problem_id = ?
            ORDER BY step_number
        """, prob["id"])

        reports.append({
            **prob,
            "components": components,
            "steps": steps
        })

    t = get_translations()
    return render_template("admin.html", 
                         reports=reports, 
                         projects=projects, 
                         managers=managers, 
                         customers=customers,
                         status_options=status,
                         active_tab=active_tab,
                         t=t)


# -------------------- ADMIN PROJECTS --------------------
@app.route("/admin/projects", methods=["GET", "POST"])
@login_required
def admin_projects():
    if request.method == "POST":
        action = request.form.get("action")

        if action == "add_project":
            project_number = request.form.get("project_number")
            project_name = request.form.get("project_name")
            manager_id = request.form.get("manager_id")
            customer_id = request.form.get("customer_id")
            quantity = request.form.get("quantity")

            # Validation
            if not project_number or not project_name or not manager_id or not customer_id or not quantity:
                flash("All fields are required.", "error")
                return redirect("/admin/projects")

            try:
                quantity = int(quantity)
            except ValueError:
                flash("Quantity must be a number.", "error")
                return redirect("/admin/projects")

            # Check if project number already exists
            existing = db.execute("SELECT id FROM projects WHERE project_number = ?", project_number)
            if existing:
                flash(f"Project number '{project_number}' already exists.", "error")
                return redirect("/admin/projects")

            # Check if project name already exists
            existing = db.execute("SELECT id FROM projects WHERE project_name = ?", project_name)
            if existing:
                flash(f"Project name '{project_name}' already exists.", "error")
                return redirect("/admin/projects")

            # Insert new project
            try:
                db.execute("""
                    INSERT INTO projects (project_number, project_name, manager_id, customer_id, quantity)
                    VALUES (?, ?, ?, ?, ?)
                """, project_number, project_name, manager_id, customer_id, quantity)
                flash(f"Project '{project_number}' added successfully!", "success")
            except Exception as e:
                flash(f"Error adding project: {str(e)}", "error")

        elif action == "delete_project":
            project_id = request.form.get("project_id")
            if project_id:
                # Check if project has associated problems
                problems = db.execute("SELECT id FROM problems WHERE project_id = ?", project_id)
                if problems:
                    flash("Cannot delete project with associated problems. Please delete problems first.", "error")
                else:
                    try:
                        db.execute("DELETE FROM projects WHERE id = ?", project_id)
                        flash("Project deleted successfully!", "success")
                    except Exception as e:
                        flash(f"Error deleting project: {str(e)}", "error")

        return redirect("/admin/projects")

    # GET: Load all projects with related data
    projects = db.execute("""
        SELECT p.id, p.project_number, p.project_name, p.quantity,
               m.manager_name, c.customer_name, c.customer_country,
               COUNT(pr.id) as problem_count
        FROM projects p
        JOIN managers m ON p.manager_id = m.id
        JOIN customers c ON p.customer_id = c.id
        LEFT JOIN problems pr ON pr.project_id = p.id
        GROUP BY p.id
        ORDER BY p.project_number
    """)

    # Get dropdown data
    managers = db.execute("SELECT id, manager_name FROM managers ORDER BY manager_name")
    customers = db.execute("SELECT id, customer_name, customer_country FROM customers ORDER BY customer_name")

    t = get_translations()
    return render_template("admin_projects.html", projects=projects, managers=managers, customers=customers, t=t)



# -------------------- RUN APP --------------------
if __name__ == "__main__":
    app.run(debug=True)
